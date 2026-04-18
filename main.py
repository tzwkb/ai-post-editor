"""
main.py — UI + orchestration for AI MTPE tool.
直接运行: python main.py
"""

import re
import sys
import datetime
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

import config
import engine
from engine import (
    get_file_headers, load_bilingual_file, build_tm_index,
    TFIDFIndex, EmbeddingIndex,
    process_file, reset_client, RICH_TEXT_AVAILABLE, _merge,
)

# ---------------------------------------------------------------------------
# UI helpers
# ---------------------------------------------------------------------------

W = 58

def banner():
    print("\n" + "╔" + "═" * W + "╗")
    print("║" + "AI Post-Edit Tool  —  Game Localization MTPE".center(W) + "║")
    print("╚" + "═" * W + "╝")

def section(title: str, step: str = ""):
    label = f"  {step}  " if step else "  "
    print(f"\n{label}{title}")
    print("  " + "─" * (W - 2))

def prompt(msg: str) -> str:
    return input(f"\n  » {msg}: ").strip()

def _strip_quotes(s: str) -> str:
    return s.strip('"').strip("'")

def pause():
    input("\n  按 Enter 继续...")

def masked_key(key: str) -> str:
    if not key:
        return "（未设置）"
    return key[:4] + "***" + key[-4:] if len(key) > 8 else "*" * len(key)


# ---------------------------------------------------------------------------
# Step 1 — API settings
# ---------------------------------------------------------------------------

def step_api_settings() -> bool:
    while True:
        section("API 配置", "【第 1 步】")
        print(f"    接口地址 : {config.API_BASE_URL}")
        print(f"    模型     : {config.MODEL}")
        print(f"    API Key  : {masked_key(config.API_KEY)}")
        print()
        print("    1. 使用以上配置，继续")
        print("    2. 修改接口地址")
        print("    3. 修改模型")
        print("    4. 修改 API Key")
        print("    0. 退出")

        choice = prompt("请选择").lower()
        if choice == "0":
            return False
        if choice == "1":
            if not config.API_KEY:
                print("\n  ⚠ API Key 未设置。")
                pause(); continue
            return True
        if choice == "2":
            v = prompt(f"接口地址（当前: {config.API_BASE_URL}，回车保持）")
            if v:
                config.API_BASE_URL = v
                reset_client()
        elif choice == "3":
            v = prompt(f"模型名称（当前: {config.MODEL}，回车保持）")
            if v:
                config.MODEL = v
        elif choice == "4":
            v = prompt("API Key（回车保持）")
            if v:
                config.API_KEY = v
                reset_client()
        else:
            print("  无效选项。")


# ---------------------------------------------------------------------------
# Step 2 — File selection
# ---------------------------------------------------------------------------

def step_select_files() -> list:
    search_dir = config.INPUT_DIR if config.INPUT_DIR.is_dir() else Path.cwd()

    while True:
        xlsx_files = sorted(search_dir.glob("*.xlsx"))
        selected: set = set()

        while True:
            section("选择文件", "【第 2 步】")
            print(f"    目录: {search_dir}\n")
            if not xlsx_files:
                print("    （未找到 .xlsx 文件）\n")
            else:
                for i, f in enumerate(xlsx_files):
                    print(f"    {'☑' if i in selected else '☐'} {i+1}. {f.name}")
            print()
            print("    编号  — 切换选中  |  A — 全选/全不选")
            print("    P     — 输入路径  |  D — 切换目录")
            print("    0     — 返回      |  Enter — 确认")

            raw = prompt("请操作")
            if raw == "":
                if not selected:
                    print("\n  ⚠ 请至少选择一个文件。"); pause(); continue
                return [str(xlsx_files[i]) for i in sorted(selected)]
            if raw == "0":
                return []
            if raw.upper() == "A":
                selected = set() if len(selected) == len(xlsx_files) else set(range(len(xlsx_files)))
                continue
            if raw.upper() == "P":
                p = _strip_quotes(prompt("文件路径"))
                if p and Path(p).is_file():
                    return [p]
                print("\n  ⚠ 文件不存在。"); pause(); continue
            if raw.upper() == "D":
                nd = _strip_quotes(prompt("新目录路径"))
                if Path(nd).is_dir():
                    search_dir = Path(nd); selected.clear(); break
                print("\n  ⚠ 目录不存在。"); pause(); continue

            # Number tokens
            tokens = re.split(r"[\s,，]+", raw)
            ok = True
            for tok in tokens:
                if tok.isdigit():
                    idx = int(tok) - 1
                    if 0 <= idx < len(xlsx_files):
                        selected.symmetric_difference_update({idx})
                    else:
                        print(f"\n  ⚠ 编号 {tok} 超出范围。"); ok = False; break
                elif tok:
                    print(f"\n  ⚠ 无法识别: {tok}"); ok = False; break
            if not ok:
                pause()


# ---------------------------------------------------------------------------
# Step 3 — Column config
# ---------------------------------------------------------------------------

def step_column_config(files: list) -> dict | None:
    ref_file = files[0]

    while True:
        section("列配置", "【第 3 步】")
        print(f"    参考文件：{Path(ref_file).name}\n")

        try:
            headers = get_file_headers(ref_file)
        except Exception as exc:
            print(f"    ⚠ 无法读取列头：{exc}"); pause(); return None

        if not headers:
            print("    ⚠ 文件第一行为空。"); pause(); return None

        print("    检测到以下列：")
        for col_idx, name in headers:
            print(f"      {col_idx}. {name}")

        col_map  = {str(c): c for c, _ in headers}
        col_name = {c: n for c, n in headers}

        def ask_col(label, required=False):
            while True:
                raw = prompt(label).strip()
                if not raw and not required:
                    return None
                if raw in col_map:
                    return col_map[raw]
                print(f"    ⚠ 请输入有效编号（{', '.join(col_map)}）。")

        print()
        tt_col  = ask_col("待审校列（TT，必选）", required=True)
        st_col  = ask_col("源语列（ST，可选）")
        key_col = ask_col("Key 列（可选）")

        ref_cols = []
        print("\n    参考列（可选，直接回车结束）：")
        while True:
            raw = prompt("添加参考列编号").strip()
            if not raw:
                break
            if raw not in col_map:
                print(f"    ⚠ 无效编号。"); continue
            rc_idx  = col_map[raw]
            rc_name = col_name.get(rc_idx, f"列{rc_idx}")
            note    = prompt(f"为「{rc_name}」提供注释").strip() or rc_name
            ref_cols.append({"col": rc_idx, "note": note})
            print(f"    ✓ 已添加：列 {rc_idx}「{rc_name}」— {note}")

        def fmt(idx):
            return f"{idx}. {col_name.get(idx,'?')}" if idx else "（不使用）"

        print(f"\n    配置预览：")
        print(f"      TT : {fmt(tt_col)}")
        print(f"      ST : {fmt(st_col)}")
        print(f"      Key: {fmt(key_col)}")
        for rc in ref_cols:
            print(f"      参考: {fmt(rc['col'])}  注释：{rc['note']}")
        print()
        print("    1. 确认  |  2. 重新选择  |  0. 返回")

        choice = prompt("请选择")
        if choice == "1":
            return {"tt": tt_col, "st": st_col, "key": key_col, "ref_cols": ref_cols}
        if choice == "0":
            return None


# ---------------------------------------------------------------------------
# Shared: load bilingual file with UI (used by step 4 & 5)
# ---------------------------------------------------------------------------

def _load_bilingual_ui(title: str, scan_dir: Path = None,
                       return_path: bool = False) -> list | None:
    """
    Shared UI for loading a bilingual xlsx/csv file (glossary or TM).
    If scan_dir is given, lists its files for quick selection.
    Returns list of (st, tt) tuples, or None to go back.
    """
    if scan_dir and scan_dir.is_dir():
        files = sorted(scan_dir.glob("*.xlsx")) + sorted(scan_dir.glob("*.csv"))
        if files:
            print(f"\n  目录：{scan_dir}")
            for i, f in enumerate(files):
                print(f"    {i+1}. {f.name}")
            raw = prompt("选择编号，或直接输入路径").strip()
            if raw.isdigit():
                idx = int(raw) - 1
                if 0 <= idx < len(files):
                    filepath = str(files[idx])
                else:
                    print("  ⚠ 编号超出范围。"); pause(); return None
            else:
                filepath = _strip_quotes(raw)
        else:
            print(f"\n  （{scan_dir} 中未找到文件）")
            filepath = _strip_quotes(prompt(f"{title}路径"))
    else:
        filepath = _strip_quotes(prompt(f"{title}路径（可直接拖入窗口）"))
    if not filepath or not Path(filepath).is_file():
        print("\n  ⚠ 文件不存在。"); pause(); return None

    try:
        headers = get_file_headers(filepath)
    except Exception as exc:
        print(f"\n  ⚠ 无法读取文件：{exc}"); pause(); return None

    print(f"\n  文件：{Path(filepath).name}")
    print("  检测到以下列：")
    for i, h in headers:
        print(f"    {i}. {h}")

    try:
        st_col = int(prompt("ST（源语）列编号"))
        tt_col = int(prompt("TT（目标语）列编号"))
        if not (1 <= st_col <= len(headers) and 1 <= tt_col <= len(headers)):
            raise ValueError
    except ValueError:
        print("\n  ⚠ 编号无效。"); pause(); return None

    try:
        entries = load_bilingual_file(filepath, st_col, tt_col)
        return (entries, filepath) if return_path else entries
    except Exception as exc:
        print(f"\n  ⚠ 加载失败：{exc}"); pause(); return None


# ---------------------------------------------------------------------------
# Step 4 — Glossary
# ---------------------------------------------------------------------------

def step_glossary() -> list | None:
    while True:
        section("术语表（可选）", "【第 4 步】")
        print("    1. 载入术语表（.xlsx / .csv）")
        print("    2. 跳过")
        print("    0. 返回")

        choice = prompt("请选择")
        if choice == "0": return None
        if choice == "2": return []
        if choice != "1": print("  无效选项。"); continue

        entries = _load_bilingual_ui("术语表文件", config.TB_DIR)
        if entries is None:
            continue
        if not entries:
            print("\n  ⚠ 未读取到有效术语。"); pause(); continue

        print(f"\n  已载入 {len(entries)} 条术语。预览（前 5 条）：")
        for st, tt in entries[:5]:
            print(f"    {st}  →  {tt}")
        print()
        print("    1. 确认  |  2. 重新选择  |  0. 返回")
        inner = prompt("请选择")
        if inner == "1": return entries
        if inner == "0": return None


# ---------------------------------------------------------------------------
# Step 5 — Translation Memory
# ---------------------------------------------------------------------------

def step_tm():
    while True:
        section("翻译记忆 TM（可选）", "【第 5 步】")
        print("    语义向量匹配（无 sentence_transformers 时自动降级为 TF-IDF）。")
        print("    1. 载入 TM 文件（.xlsx / .csv）")
        print("    2. 跳过")
        print("    0. 返回")

        choice = prompt("请选择")
        if choice == "0": return None
        if choice == "2": return []
        if choice != "1": print("  无效选项。"); continue

        result = _load_bilingual_ui("TM 文件", config.TM_DIR, return_path=True)
        if result is None:
            continue
        entries, tm_filepath = result
        if not entries:
            print("\n  ⚠ 未读取到有效条目。"); pause(); continue

        print(f"\n  已载入 {len(entries)} 条，正在构建索引...")
        try:
            idx, desc = build_tm_index(entries, source_path=tm_filepath)
        except Exception as exc:
            print(f"\n  ⚠ 索引构建失败：{exc}"); pause(); continue
        print(f"  ✓ {desc}")

        sample = entries[min(5, len(entries) - 1)][0]
        hits = idx.query(sample, top_k=3)
        if hits:
            print(f"\n  示例查询：「{sample[:50]}」")
            for score, st, tt in hits:
                print(f"    [{score:.0%}] {st[:40]} → {tt[:40]}")

        print()
        print("    1. 确认  |  2. 重新选择  |  0. 返回")
        inner = prompt("请选择")
        if inner == "1": return idx
        if inner == "0": return None


# ---------------------------------------------------------------------------
# Step 6 — Confirm and run
# ---------------------------------------------------------------------------

def step_confirm(files: list, col_config: dict, glossary: list, tm_index) -> bool | None:
    section("确认并开始处理", "【第 6 步】")
    print(f"    接口地址 : {config.API_BASE_URL}")
    print(f"    模型     : {config.MODEL}")
    print(f"    术语表   : {f'{len(glossary)} 条' if glossary else '未使用'}")
    if tm_index:
        idx_type = "语义向量" if isinstance(tm_index, EmbeddingIndex) else "TF-IDF"
        print(f"    翻译记忆 : {len(tm_index.entries)} 条（{idx_type}）")
    else:
        print(f"    翻译记忆 : 未使用")
    print(f"    列配置   : TT={col_config['tt']}  ST={col_config.get('st') or '—'}  Key={col_config.get('key') or '—'}")
    for rc in col_config.get("ref_cols", []):
        print(f"    参考列   : 列 {rc['col']}  注释：{rc['note']}")
    print(f"    待处理   : {len(files)} 个文件")
    for f in files:
        print(f"      • {Path(f).name}")
    print()
    print("    输出：原文件名 + _postedit.xlsx")
    if not RICH_TEXT_AVAILABLE:
        print("    ⚠ openpyxl 版本过低，加粗标记将以纯文本输出")
    print()
    print("    1. 开始  |  2. 返回  |  0. 退出")

    while True:
        choice = prompt("请选择")
        if choice == "1": return True
        if choice == "2": return None
        if choice == "0": return False
        print("  无效选项。")


# ---------------------------------------------------------------------------
# Step 7 — QA (optional)
# ---------------------------------------------------------------------------

def step_qa(files: list, col_config: dict, glossary: list, tm_index) -> dict | None:
    while True:
        section("AI QA 审查（可选）", "【第 7 步】")
        print("    L1 — 规则检查（无需 API）")
        print("    L2 — AI 语义 QA（整 sheet 批量发送）")
        print("    失败行自动修复（最多 2 次迭代）")
        print("    结果写入 'QA Status' / 'QA Fixed' 列\n")
        print("    1. 运行 QA  |  2. 跳过  |  0. 退出")

        choice = prompt("请选择")
        if choice in ("0", "2"): return None
        if choice != "1": print("  无效选项。"); continue

        try:
            import qa_module
        except ImportError as exc:
            print(f"\n  ⚠ 无法载入 qa_module.py：{exc}"); pause(); return None

        global_qa = {"pass": 0, "total": 0, "by_type": Counter(), "fail_rows": []}

        for filepath in files:
            p = Path(filepath)
            out_path = config.OUTPUT_DIR / (p.stem + config.OUTPUT_SUFFIX + p.suffix)
            if not out_path.exists():
                print(f"  ⚠ 找不到输出文件：{out_path.name}，跳过。"); continue

            print(f"\n  QA: {out_path.name}")
            wb = load_workbook(str(out_path), rich_text=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"  Sheet: {sheet_name}")

                # Find Post Edit EN column
                pe_col = None
                for r in range(1, 4):
                    for c in range(1, ws.max_column + 1):
                        if ws.cell(row=r, column=c).value == config.OUTPUT_COLUMN_HEADER:
                            pe_col = c; break
                    if pe_col: break

                if pe_col is None:
                    print(f"    ⚠ 找不到 '{config.OUTPUT_COLUMN_HEADER}' 列，跳过。"); continue

                summary = qa_module.qa_and_repair(
                    sheet=ws, post_edit_col=pe_col,
                    col_config=col_config, glossary=glossary, tm_index=tm_index,
                )
                global_qa["pass"]  += summary["pass"]
                global_qa["total"] += summary["total"]
                global_qa["by_type"].update(summary["by_type"])
                global_qa["fail_rows"].extend(summary["fail_rows"])

            # Save
            try:
                wb.save(str(out_path))
                print(f"  Saved → {out_path.name}")
            except PermissionError:
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                fb = out_path.parent / (out_path.stem + f"_qa_{ts}" + out_path.suffix)
                wb.save(str(fb))
                print(f"  [WARNING] File locked — saved to {fb.name}")

        t, p = global_qa["total"], global_qa["pass"]
        print(f"\n  {'─'*50}")
        print(f"  QA 汇总：{p}/{t} PASS，{t - p} 行有问题")
        if global_qa["by_type"]:
            for etype, cnt in global_qa["by_type"].most_common():
                print(f"    {etype}: {cnt}")
        print(f"  {'─'*50}")
        return global_qa


# ---------------------------------------------------------------------------
# Run log
# ---------------------------------------------------------------------------

def write_run_log(files: list, qa_summary: dict | None, note: str = ""):
    import openpyxl as _opx
    log_path = config.BASE_DIR / "run_log.xlsx"
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if log_path.exists():
        try:
            wb = load_workbook(str(log_path))
        except Exception:
            wb = _opx.Workbook()
            wb.active.append(["序号", "时间", "测试文件", "AIQA结果", "备注"])
        ws = wb.active
        seq = ws.max_row
    else:
        wb = _opx.Workbook()
        ws = wb.active
        ws.append(["序号", "时间", "测试文件", "AIQA结果", "备注"])
        seq = 1

    file_names = "\n".join(Path(f).name for f in files)

    if qa_summary and qa_summary.get("total", 0) > 0:
        p, t = qa_summary["pass"], qa_summary["total"]
        qa_str = f"PASS {p}/{t}"
        if qa_summary.get("by_type"):
            qa_str += "  [" + ", ".join(
                f"{k}:{v}" for k, v in
                sorted(qa_summary["by_type"].items(), key=lambda x: x[1], reverse=True)
            ) + "]"
    else:
        qa_str = "（未运行 QA）"

    ws.append([seq, now, file_names, qa_str, note or ""])
    try:
        wb.save(str(log_path))
        print(f"\n  [Log] 已写入 run_log.xlsx（第 {seq} 行）")
    except PermissionError:
        print("\n  ⚠ run_log.xlsx 被占用，无法写入。")


# ---------------------------------------------------------------------------
# main()
# ---------------------------------------------------------------------------

def main():
    banner()

    # Ensure all working directories exist
    for d in [config.INPUT_DIR, config.OUTPUT_DIR, config.TM_DIR, config.TB_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    if not step_api_settings():
        print("\n  已退出。\n"); return

    while True:
        files = step_select_files()
        if not files:
            if not step_api_settings():
                print("\n  已退出。\n"); return
            continue

        col_config = step_column_config(files)
        if col_config is None: continue

        glossary = step_glossary()
        if glossary is None: continue

        tm_result = step_tm()
        if tm_result is None: continue
        tm_index = tm_result if isinstance(tm_result, (TFIDFIndex, EmbeddingIndex)) else None

        result = step_confirm(files, col_config, glossary, tm_index)
        if result is True: break
        if result is None: continue
        print("\n  已退出。\n"); return

    # Process
    print()
    succeeded, all_term_counts = 0, {}

    for filepath in files:
        ok, tc = process_file(filepath, col_config=col_config,
                              glossary=glossary, tm_index=tm_index,
                              retry_prompt=prompt)
        if ok: succeeded += 1
        _merge(all_term_counts, tc)

    # Summary
    print("\n" + "╔" + "═" * W + "╗")
    print("║" + f"处理完成：{succeeded} / {len(files)} 个文件成功".center(W) + "║")
    print("╚" + "═" * W + "╝")

    if all_term_counts:
        print("\n  术语应用统计")
        print(f"  {'─'*54}")
        print(f"  {'ST（源语）':<20}  {'TT（目标语）':<22}  命中次数")
        print(f"  {'─'*54}")
        for (st, tt), count in sorted(all_term_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {st:<20}  {tt:<22}  {count}")
        total_hits = sum(all_term_counts.values())
        print(f"  共 {len(all_term_counts)} 个术语命中，合计 {total_hits} 次\n")
    elif glossary:
        print("\n  （本次处理未命中任何术语）\n")

    # QA
    qa_summary = step_qa(files, col_config, glossary, tm_index)

    # Log
    note = prompt("本次运行备注（回车跳过）")
    write_run_log(files, qa_summary, note)

    pause()


if __name__ == "__main__":
    main()
