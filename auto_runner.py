"""
auto_runner.py — Non-interactive runner for AI MTPE tool.
All options are hardcoded. Run: python auto_runner.py
"""
import datetime
import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
import config
import engine
import qa_module

# ---------------------------------------------------------------------------
# Hardcoded configuration — edit here to change behaviour
# ---------------------------------------------------------------------------

COL_CONFIG = {
    "tt": 6,        # 英文 (English MT to post-edit)
    "st": 4,        # 中文简体 (Chinese source)
    "key": 1,       # id
    "ref_cols": [{"col": 3, "note": "DNT"}],
}

TB_FILE   = config.TB_DIR / "tb.xlsx"
TB_SHEET  = "最新术语表（4月11日更新）"
TB_ST_COL = 1   # 中文
TB_TT_COL = 3   # 英文（第3列）

TM_FILE   = config.TM_DIR / "超大TM.xlsx"
TM_ST_COL = 1   # 对白
TM_TT_COL = 2   # 译文

RUN_QA    = True

# ---------------------------------------------------------------------------

def _load_glossary(filepath, sheet_name, st_col, tt_col, skip_rows=0):
    wb = load_workbook(str(filepath), data_only=True)
    ws = wb[sheet_name]
    rows = [
        (str(row[st_col - 1] or "").strip(), str(row[tt_col - 1] or "").strip().strip('\n'))
        for i, row in enumerate(ws.iter_rows(values_only=True))
        if i >= skip_rows
        and row and len(row) >= max(st_col, tt_col)
        and str(row[st_col - 1] or "").strip()
        and str(row[tt_col - 1] or "").strip()
    ]
    wb.close()
    return rows


def _run_qa(filepath, col_config, glossary, tm_index, row_term_hits=None):
    p = Path(filepath)
    out_path = config.OUTPUT_DIR / (p.stem + config.OUTPUT_SUFFIX + p.suffix)
    if not out_path.exists():
        print(f"  QA: output not found for {p.name}, skipping.")
        return {"pass": 0, "total": 0, "by_type": Counter(), "fail_rows": []}

    if row_term_hits is None:
        row_term_hits = {}

    print(f"\n  QA: {out_path.name}")
    wb = load_workbook(str(out_path), rich_text=True)
    total_summary = {"pass": 0, "total": 0, "by_type": Counter(), "fail_rows": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name}")

        pe_col = None
        for r in range(1, 4):
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value == config.OUTPUT_COLUMN_HEADER:
                    pe_col = c
                    break
            if pe_col:
                break

        if pe_col is None:
            print(f"    No '{config.OUTPUT_COLUMN_HEADER}' column, skipping.")
            continue

        s = qa_module.qa_and_repair(
            sheet=ws, post_edit_col=pe_col,
            col_config=col_config, glossary=glossary, tm_index=tm_index,
            row_term_hits=row_term_hits,
        )
        total_summary["pass"]  += s["pass"]
        total_summary["total"] += s["total"]
        total_summary["by_type"].update(s["by_type"])
        total_summary["fail_rows"].extend(s["fail_rows"])

    try:
        wb.save(str(out_path))
        print(f"  Saved: {out_path.name}")
    except PermissionError:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fb = out_path.parent / (out_path.stem + f"_qa_{ts}" + out_path.suffix)
        wb.save(str(fb))
        print(f"  [WARNING] File locked — saved to {fb.name}")

    return total_summary


def _write_log(files, qa_summary, note="auto_runner"):
    import openpyxl as _opx
    log_path = config.BASE_DIR / "run_log.xlsx"
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if log_path.exists():
        try:
            wb = load_workbook(str(log_path))
        except Exception:
            wb = _opx.Workbook()
            wb.active.append(["序号", "时间", "测试文件", "AIQA结果", "备注"])
        ws = wb.active
        seq = ws.max_row
    else:
        wb = _opx.Workbook()
        ws = wb.active
        ws.append(["序号", "时间", "测试文件", "AIQA结果", "备注"])
        seq = 1

    file_names = "\n".join(Path(f).name for f in files)
    if qa_summary and qa_summary.get("total", 0) > 0:
        p, t = qa_summary["pass"], qa_summary["total"]
        qa_str = f"PASS {p}/{t}"
        if qa_summary.get("by_type"):
            qa_str += "  [" + ", ".join(
                f"{k}:{v}" for k, v in
                sorted(qa_summary["by_type"].items(), key=lambda x: x[1], reverse=True)
            ) + "]"
    else:
        qa_str = "(QA not run)"

    ws.append([seq, now, file_names, qa_str])
    try:
        wb.save(str(log_path))
        print(f"\n[Log] Written to run_log.xlsx (row {seq})")
    except PermissionError:
        print("\n[WARNING] run_log.xlsx is locked, could not write.")


def main():
    for d in [config.INPUT_DIR, config.OUTPUT_DIR, config.TM_DIR, config.TB_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    engine.init_api_log()

    # Input files — default to test.xlsx if it exists, otherwise all xlsx
    test_file = config.INPUT_DIR / "test.xlsx"
    if test_file.exists():
        files = [str(test_file)]
        print(f"[TEST MODE] Using test.xlsx")
    else:
        files = sorted(str(p) for p in config.INPUT_DIR.glob("*.xlsx"))
    if not files:
        print("No .xlsx files found in input/")
        return
    print(f"Input: {len(files)} file(s)")
    for f in files:
        print(f"  {Path(f).name}")

    # Glossary
    glossary = []
    if TB_FILE.exists():
        glossary = _load_glossary(TB_FILE, TB_SHEET, TB_ST_COL, TB_TT_COL, skip_rows=2)
        print(f"\nGlossary: {len(glossary)} terms  ({TB_FILE.name} / {TB_SHEET})")
    else:
        print(f"\n[WARN] TB file not found: {TB_FILE.name}")

    # TM
    tm_index = None
    if TM_FILE.exists():
        entries = engine.load_bilingual_file(str(TM_FILE), TM_ST_COL, TM_TT_COL)
        tm_index, desc = engine.build_tm_index(entries, source_path=str(TM_FILE))
        print(f"TM: {desc}")
    else:
        print(f"[WARN] TM file not found: {TM_FILE.name}")

    # MTPE
    print("\n" + "=" * 60)
    succeeded, all_term_counts, all_row_term_hits = 0, {}, {}
    for filepath in files:
        ok, tc, rth = engine.process_file(
            filepath, col_config=COL_CONFIG,
            glossary=glossary, tm_index=tm_index,
        )
        if ok:
            succeeded += 1
        engine._merge(all_term_counts, tc)
        all_row_term_hits.update(rth)

    print(f"\nMTPE done: {succeeded}/{len(files)} files OK")

    # QA — only run if all MTPE rows completed (no failures remain)
    total_mtpe_failed = sum(len(v) for v in getattr(engine, '_last_failed', {}).values()) if False else 0
    # Check by re-scanning output files for empty PE cells
    import openpyxl as _opx
    mtpe_incomplete = False
    for filepath in files:
        p = Path(filepath)
        out_path = config.OUTPUT_DIR / (p.stem + config.OUTPUT_SUFFIX + p.suffix)
        if not out_path.exists():
            mtpe_incomplete = True
            break
        wb_check = _opx.load_workbook(str(out_path), data_only=True)
        for sname in wb_check.sheetnames:
            ws_check = wb_check[sname]
            pe_col_check = None
            for r in range(1, 4):
                for c in range(1, ws_check.max_column + 1):
                    if ws_check.cell(row=r, column=c).value == config.OUTPUT_COLUMN_HEADER:
                        pe_col_check = c
                        break
                if pe_col_check:
                    break
            if pe_col_check is None:
                continue
            header_r = next((r for r in range(1, 4) if ws_check.cell(row=r, column=pe_col_check).value == config.OUTPUT_COLUMN_HEADER), 1)
            empty = sum(1 for r in range(header_r + 1, ws_check.max_row + 1)
                       if not str(ws_check.cell(row=r, column=pe_col_check).value or "").strip())
            if empty > 0:
                print(f"\n[QA SKIPPED] {out_path.name} still has {empty} empty Post Edit rows — complete MTPE first.")
                mtpe_incomplete = True
        wb_check.close()
        if mtpe_incomplete:
            break

    global_qa = {"pass": 0, "total": 0, "by_type": Counter(), "fail_rows": []}
    if mtpe_incomplete:
        _write_log(files, None)
        return

    if RUN_QA:
        for filepath in files:
            s = _run_qa(filepath, COL_CONFIG, glossary, tm_index, all_row_term_hits)
            global_qa["pass"]  += s["pass"]
            global_qa["total"] += s["total"]
            global_qa["by_type"].update(s["by_type"])
            global_qa["fail_rows"].extend(s["fail_rows"])

        t, p = global_qa["total"], global_qa["pass"]
        print(f"\nQA Summary: {p}/{t} PASS, {t - p} failing")
        for etype, cnt in global_qa["by_type"].most_common():
            print(f"  {etype}: {cnt}")

    _write_log(files, global_qa if RUN_QA else None)


if __name__ == "__main__":
    main()
