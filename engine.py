"""
engine.py — Core processing engine for AI MTPE tool.

Contains:
- Prompt loading
- API client + call helpers
- TM index (Sentence Embedding)
- Cell / file utilities
- Sheet & file processing
"""

import asyncio
import datetime
import difflib
import json
import os
import re
import time
import threading
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    RICH_TEXT_AVAILABLE = True
except ImportError:
    RICH_TEXT_AVAILABLE = False

try:
    from openai import OpenAI, AsyncOpenAI
except ImportError:
    raise ImportError("openai package not found. Run: pip install openai openpyxl")

import config

# ---------------------------------------------------------------------------
# Prompt loading
# ---------------------------------------------------------------------------

PROMPTS_DIR = Path(__file__).parent / "prompts"
_prompt_cache: dict = {}

# ---------------------------------------------------------------------------
# API call logger
# ---------------------------------------------------------------------------

_log_lock = threading.Lock()
_log_file = None   # set once per run by init_api_log()

def init_api_log():
    """
    Open (or create) today's API log file.
    Must be called once at program start before any API calls.
    Path: logs/api_YYYYMMDD_HHMMSS.jsonl
    """
    global _log_file
    log_dir = Path(__file__).parent / "logs"
    log_dir.mkdir(exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = log_dir / f"api_{ts}.jsonl"
    _log_file = open(log_path, "a", encoding="utf-8", buffering=1)
    print(f"  [Log] API log -> logs/api_{ts}.jsonl")
    return log_path


def _log_api_call(messages: list, response: str | None, error: str | None = None):
    """Append one JSONL record with request messages and response."""
    if _log_file is None:
        return
    record = {
        "ts": datetime.datetime.now().isoformat(timespec="seconds"),
        "model": config.MODEL,
        "messages": messages,
        "response": response,
    }
    if error:
        record["error"] = error
    with _log_lock:
        _log_file.write(json.dumps(record, ensure_ascii=False) + "\n")

# Active project prompt file is configured in config.MTPE_PROJECT_FILE.

def load_prompt(name: str) -> str:
    """
    Load a prompt by name (cached after first read).

    For name="mtpe": concatenates config.MTPE_PROJECT_FILE + mtpe_base.md so
    the project context always precedes the universal rules.
    For all other names: loads prompts/<name>.md directly.
    """
    if name not in _prompt_cache:
        if name == "mtpe":
            project = (PROMPTS_DIR / config.MTPE_PROJECT_FILE).read_text(encoding="utf-8")
            base    = (PROMPTS_DIR / "mtpe_base.md").read_text(encoding="utf-8")
            _prompt_cache[name] = project.rstrip() + "\n\n" + base
        else:
            p = PROMPTS_DIR / f"{name}.md"
            if not p.exists():
                raise FileNotFoundError(f"Prompt file not found: {p}")
            _prompt_cache[name] = p.read_text(encoding="utf-8")
    return _prompt_cache[name]


# ---------------------------------------------------------------------------
# API clients (sync for QA repair loop, async for MTPE hot path)
# ---------------------------------------------------------------------------

_client: OpenAI | None = None
_async_client: AsyncOpenAI | None = None

def get_client() -> OpenAI:
    global _client
    if _client is None:
        _client = OpenAI(api_key=config.API_KEY, base_url=config.API_BASE_URL)
    return _client

def get_async_client() -> AsyncOpenAI:
    global _async_client
    if _async_client is None:
        _async_client = AsyncOpenAI(api_key=config.API_KEY, base_url=config.API_BASE_URL)
    return _async_client

def reset_client():
    """Force re-create both clients (call after config changes)."""
    global _client, _async_client
    _client = OpenAI(api_key=config.API_KEY, base_url=config.API_BASE_URL)
    _async_client = AsyncOpenAI(api_key=config.API_KEY, base_url=config.API_BASE_URL)


def _call_api_raw(messages: list, temperature: float = None,
                  attempt: int = 0) -> str | None:
    """
    Synchronous API call with exponential-backoff retry.
    Used by QA repair loop (runs in a ThreadPoolExecutor inside an async context).
    """
    temp = temperature if temperature is not None else config.TEMPERATURE
    try:
        resp = get_client().chat.completions.create(
            model=config.MODEL,
            messages=messages,
            max_tokens=config.MAX_TOKENS,
            temperature=temp,
            timeout=config.REQUEST_TIMEOUT,
        )
        text = resp.choices[0].message.content.strip()
        if attempt == 0:
            _log_api_call(messages, text)
        return text
    except Exception as exc:
        if attempt < config.MAX_RETRIES:
            wait = config.RETRY_DELAY * (2 ** attempt)
            print(f"\n    [Retry {attempt+1}/{config.MAX_RETRIES}] {exc} — waiting {wait}s...")
            time.sleep(wait)
            return _call_api_raw(messages, temp, attempt + 1)
        _log_api_call(messages, None, error=str(exc))
        print(f"\n    [FAILED] {exc}")
        return None


async def _async_call_api_raw(messages: list, temperature: float = None,
                               semaphore: asyncio.Semaphore = None,
                               attempt: int = 0) -> str | None:
    """
    Async API call with exponential-backoff retry.
    Used by process_sheet hot path — all concurrent requests share one event loop.
    semaphore: controls max concurrent in-flight requests.
    """
    temp = temperature if temperature is not None else config.TEMPERATURE
    async with (semaphore or _NULL_SEMAPHORE):
        try:
            resp = await get_async_client().chat.completions.create(
                model=config.MODEL,
                messages=messages,
                max_tokens=config.MAX_TOKENS,
                temperature=temp,
                timeout=config.REQUEST_TIMEOUT,
            )
            text = resp.choices[0].message.content.strip()
            if attempt == 0:
                _log_api_call(messages, text)
            return text
        except Exception as exc:
            if attempt < config.MAX_RETRIES:
                wait = config.RETRY_DELAY * (2 ** attempt)
                print(f"\n    [Retry {attempt+1}/{config.MAX_RETRIES}] {exc} — waiting {wait}s...")
                await asyncio.sleep(wait)
                return await _async_call_api_raw(messages, temp, semaphore, attempt + 1)
            _log_api_call(messages, None, error=str(exc))
            print(f"\n    [FAILED] {exc}")
            return None


class _null_semaphore:
    """No-op async context manager — singleton used when no semaphore is passed."""
    async def __aenter__(self): return self
    async def __aexit__(self, *_): pass

_NULL_SEMAPHORE = _null_semaphore()



def strip_json_fences(text: str) -> str:
    """Strip markdown code fences (```json ... ```) from AI JSON responses."""
    text = re.sub(r'^```[a-z]*\n?', '', text.strip())
    return re.sub(r'\n?```$', '', text).strip()


def _sanitize(text: str) -> str:
    """
    Undo pipe-character escapes and normalise [[B]]/[[/B]] bold markup.

    Bold normalisation handles common AI output errors:
    1. Single-bracket close tag  [/B]  -> [[/B]]
    2. Inverted open/close order ([[/B]]...[[B]]) -> strip both
    3. Orphan tags (open with no close, or close with no open) -> strip
    After normalisation, malformed pairs no longer reach parse_bold_markup().
    """
    if not text:
        return text

    # Pipe escapes
    text = re.sub(r'__PIPE__|<PIPE>|\[PIPE\]|\\\|', '|', text)

    # 1. Normalise [/B] (single-bracket close) -> [[/B]]
    #    Must NOT already be [[/B]] — use negative lookbehind
    text = re.sub(r'(?<!\[)\[/B\](?!\])', '[[/B]]', text)

    # 2. Strip inverted pairs: [[/B]] ... [[B]] (close before open)
    def _strip_inverted(t: str) -> str:
        # Find first [[/B]] that appears before any [[B]]
        while True:
            close = t.find('[[/B]]')
            open_ = t.find('[[B]]')
            if close == -1 or (open_ != -1 and open_ < close):
                break
            # Close tag found before open — strip it
            t = t[:close] + t[close + 6:]
        return t
    text = _strip_inverted(text)

    # 3. Ensure open/close counts match; strip orphans
    opens  = text.count('[[B]]')
    closes = text.count('[[/B]]')
    if opens != closes:
        # Strip all bold markup — mismatched tags are unsafe to render
        text = re.sub(r'\[\[/?B\]\]', '', text)

    # 4. Replace em-dashes with contextually appropriate punctuation
    #    Em-dashes are forbidden in both dialogue and narration.
    #    Space-padded dash ( — ) -> comma; bare dash (—/–) at word boundary -> comma.
    text = re.sub(r'\s*[–—]\s*', ', ', text)

    # 4b. "Narration" is a speaker label in this project — never change it.
    #     AI sometimes "corrects" it to "Narrator"; revert that here.
    text = re.sub(r'\bNarrator\b', 'Narration', text)
    text = re.sub(r'\bnarrator\b', 'narration', text)

    # 5. Auto-close dangling quotes at the start/end of each logical line.
    #    Only fixes lines where exactly one side has the quote marker.
    def _fix_dangling_quotes(t: str) -> str:
        fixed = []
        for line in t.split('\n'):
            s = line.strip()
            if not s:
                fixed.append(line)
                continue
            # Double quotes: start has " but end doesn't (and vice-versa)
            has_open  = s[0] == '"'
            has_close = s[-1] == '"'
            if has_open and not has_close:
                line = line.rstrip() + '"'
            elif has_close and not has_open:
                line = '"' + line.lstrip()
            # Single quotes (skip contractions: length > 3 avoids "it's" false positives)
            s = line.strip()  # re-strip after possible modification
            has_open_sq  = s[0] == "'"
            has_close_sq = s[-1] == "'"
            if has_open_sq and not has_close_sq and len(s) > 3:
                line = line.rstrip() + "'"
            elif has_close_sq and not has_open_sq and len(s) > 3:
                line = "'" + line.lstrip()
            fixed.append(line)
        return '\n'.join(fixed)
    text = _fix_dangling_quotes(text)

    return text


# ---------------------------------------------------------------------------
# MTPE API calls
# ---------------------------------------------------------------------------

def _build_mtpe_messages(source_text: str, en_text: str, key: str = "",
                         glossary_hints: list = None, ref_data: list = None,
                         tm_refs: list = None) -> list:
    """Build the messages list for a single MTPE request."""
    parts = []
    if key:
        parts.append(f"Context key: {key}")
    if glossary_hints:
        lines = "\n".join(f"  - {s} → {t}" for s, t in glossary_hints)
        parts.append(f"Mandatory terminology (MUST use these exact translations):\n{lines}")
    if tm_refs:
        lines = "\n".join(f"  - [{sc:.0%}] {s} → {t}" for sc, s, t in tm_refs)
        parts.append(f"Style references from Translation Memory (tone/style guide ONLY):\n{lines}")
    if ref_data:
        lines = "\n".join(f"  [{d['note']}]: {d['value']}" for d in ref_data)
        parts.append(f"Reference information:\n{lines}")
    if source_text and source_text.strip():
        parts.append(f"SOURCE (Chinese):\n{source_text}")
    parts.append(f"TRANSLATION (English to post-edit):\n{en_text}")

    return [
        {"role": "system", "content": load_prompt("mtpe")},
        {"role": "user",   "content": "\n\n".join(parts)},
    ]


def call_api(source_text: str, en_text: str, key: str = "",
             glossary_hints: list = None, ref_data: list = None,
             tm_refs: list = None) -> str | None:
    """Synchronous post-edit — used by QA repair loop."""
    if not en_text or not en_text.strip():
        return en_text
    msgs = _build_mtpe_messages(source_text, en_text, key,
                                glossary_hints, ref_data, tm_refs)
    result = _call_api_raw(msgs)
    return _sanitize(result) if result else None


def _build_cluster_messages(st_text: str, tt_variants: list, key_texts: list,
                            glossary_hints: list = None, ref_data: list = None,
                            tm_refs: list = None) -> list:
    """Build messages list for a multi-variant cluster MTPE request."""
    parts = []
    unique_keys = list(dict.fromkeys(k for k in key_texts if k))
    if unique_keys:
        parts.append("Context key(s): " + " | ".join(unique_keys))
    if glossary_hints:
        lines = "\n".join(f"  - {s} → {t}" for s, t in glossary_hints)
        parts.append(f"Mandatory terminology (MUST use these exact translations):\n{lines}")
    if tm_refs:
        lines = "\n".join(f"  [{i+1}] {s} → {t}" for i, (_, s, t) in enumerate(tm_refs))
        parts.append(f"Style reference from Translation Memory (tone/style guide ONLY):\n{lines}")
    if ref_data:
        lines = "\n".join(f"  [{d['note']}]: {d['value']}" for d in ref_data)
        parts.append(f"Reference information:\n{lines}")
    if st_text and st_text.strip():
        parts.append(f"SOURCE (Chinese):\n{st_text}")
    variants_block = "\n".join(f"  [{i+1}] {v}" for i, v in enumerate(tt_variants))
    parts.append(
        f"The following {len(tt_variants)} translations are all for the same source text.\n"
        f"Compare them, select the best one as a base, then post-edit it.\n"
        f"Output ONLY the final post-edited result — no labels, no explanations.\n\n"
        f"TRANSLATION VARIANTS:\n{variants_block}"
    )
    return [
        {"role": "system", "content": load_prompt("mtpe")},
        {"role": "user",   "content": "\n\n".join(parts)},
    ]


def call_api_cluster(st_text: str, tt_variants: list, key_texts: list,
                     glossary_hints: list = None, ref_data: list = None,
                     tm_refs: list = None) -> str | None:
    """Synchronous cluster post-edit — used by QA repair loop."""
    if not tt_variants:
        return None
    if len(tt_variants) == 1:
        return call_api(st_text, tt_variants[0],
                        key_texts[0] if key_texts else "",
                        glossary_hints, ref_data, tm_refs)
    msgs = _build_cluster_messages(st_text, tt_variants, key_texts,
                                   glossary_hints, ref_data, tm_refs)
    result = _call_api_raw(msgs)
    return _sanitize(result) if result else None


async def _async_call_api(source_text: str, en_text: str, key: str = "",
                          glossary_hints: list = None, ref_data: list = None,
                          tm_refs: list = None,
                          semaphore: asyncio.Semaphore = None) -> str | None:
    """Async post-edit — used by process_sheet hot path."""
    if not en_text or not en_text.strip():
        return en_text
    msgs = _build_mtpe_messages(source_text, en_text, key,
                                glossary_hints, ref_data, tm_refs)
    result = await _async_call_api_raw(msgs, semaphore=semaphore)
    return _sanitize(result) if result else None


async def _async_call_api_cluster(st_text: str, tt_variants: list, key_texts: list,
                                  glossary_hints: list = None, ref_data: list = None,
                                  tm_refs: list = None,
                                  semaphore: asyncio.Semaphore = None) -> str | None:
    """Async cluster post-edit — used by process_sheet hot path."""
    if not tt_variants:
        return None
    if len(tt_variants) == 1:
        return await _async_call_api(st_text, tt_variants[0],
                                     key_texts[0] if key_texts else "",
                                     glossary_hints, ref_data, tm_refs, semaphore)
    msgs = _build_cluster_messages(st_text, tt_variants, key_texts,
                                   glossary_hints, ref_data, tm_refs)
    result = await _async_call_api_raw(msgs, semaphore=semaphore)
    return _sanitize(result) if result else None


# ---------------------------------------------------------------------------
# Batch MTPE API calls
# ---------------------------------------------------------------------------

def _build_batch_messages(batch_items: list) -> list:
    """
    Build messages for a batch MTPE request.

    batch_items: list of dicts with keys:
        label      — int index (1-based) used to label each row in prompt/response
        st         — source Chinese text
        tt         — English MT text (single string, already cluster-resolved)
        key        — context key (may be empty)
        hints      — list of (st, tt) glossary pairs
        tm_hits    — list of (score, st, tt) TM references
        ref_data   — list of {note, value} dicts (DNT etc.)

    AI must return results as:
        [1]
        <post-edited text>
        [2]
        <post-edited text>
        ...
    """
    system_content = load_prompt("mtpe")

    # Collect shared glossary terms across all items in batch
    all_hints = {}
    for item in batch_items:
        for s, t in (item.get("hints") or []):
            all_hints[s] = t

    preamble_parts = []
    if all_hints:
        lines = "\n".join(f"  - {s} → {t}" for s, t in all_hints.items())
        preamble_parts.append(f"Mandatory terminology (MUST use these exact translations):\n{lines}")

    # Build per-row blocks
    row_blocks = []
    for item in batch_items:
        label = item["label"]
        st    = item["st"]
        tt    = item["tt"]
        key   = item.get("key", "")
        tm_hits = item.get("tm_hits") or []
        ref_data = item.get("ref_data") or []

        block_parts = []
        if key:
            block_parts.append(f"Key: {key}")
        if tm_hits:
            tm_lines = "\n".join(f"    [{sc:.0%}] {s} → {t}" for sc, s, t in tm_hits)
            block_parts.append(f"TM references (style guide only):\n{tm_lines}")
        for rd in ref_data:
            block_parts.append(f"[{rd['note']}]: {rd['value']}")
        block_parts.append(f"SOURCE: {st.replace(chr(10), ' / ')}")
        block_parts.append(f"TRANSLATION: {tt.replace(chr(10), ' / ')}")

        row_blocks.append(f"[{label}]\n" + "\n".join(block_parts))

    instructions = (
        f"Post-edit the following {len(batch_items)} rows.\n"
        "Return a JSON object mapping each row number (as a string key) to its post-edited result.\n"
        'Example for 3 rows: {"1": "Post-edited text...", "2": "...", "3": "..."}\n\n'
        "Rules:\n"
        "- Apply ALL rules from your system prompt.\n"
        "- Mark changes with [[B]]...[[/B]] tags as usual.\n"
        "- Output ONLY valid JSON — no markdown fences, no explanations.\n"
        "- If a row needs no changes, output it exactly as the TRANSLATION.\n"
        "- Preserve all game tags ({b}, {/b}, {nl}, etc.) exactly.\n"
        "- Preserve \\n sequences as literal backslash-n in the JSON string value.\n"
    )

    user_parts = []
    if preamble_parts:
        user_parts.extend(preamble_parts)
    user_parts.append(instructions)
    user_parts.append("--- ROWS ---")
    user_parts.extend(row_blocks)
    user_parts.append("--- END ---")

    return [
        {"role": "system", "content": system_content},
        {"role": "user",   "content": "\n\n".join(user_parts)},
    ]


def _parse_batch_response(raw: str, labels: list) -> dict:
    """
    Parse AI batch response into {label: result_text}.
    Expected format: JSON object {"1": "text", "2": "text", ...}
    Falls back to [N]\\ntext pattern if JSON parsing fails.

    Returns dict mapping label (int) -> post-edited string.
    Missing labels map to None.
    """
    if not raw:
        return {lbl: None for lbl in labels}

    label_set = set(labels)
    result = {}

    # Primary: JSON parsing
    try:
        data = json.loads(strip_json_fences(raw))
        if isinstance(data, dict):
            for k, v in data.items():
                try:
                    lbl = int(k)
                    if lbl in label_set and isinstance(v, str):
                        result[lbl] = v
                except (ValueError, TypeError):
                    pass
    except (json.JSONDecodeError, ValueError, AttributeError):
        pass

    # Fallback: [N]\ntext pattern
    if not result:
        parts = re.split(r'(?m)^\[(\d+)\]\s*\n?', raw)
        i = 1
        while i + 1 < len(parts):
            try:
                lbl = int(parts[i])
                content = parts[i + 1].strip()
                if lbl in label_set:
                    result[lbl] = content
            except (ValueError, IndexError):
                pass
            i += 2

    # Fill missing labels with None
    for lbl in labels:
        if lbl not in result:
            result[lbl] = None

    return result


async def _async_call_api_batch(batch_items: list,
                                 semaphore: asyncio.Semaphore = None) -> dict:
    """
    Async batch MTPE call. Returns dict of label -> post-edited string (or None on failure).
    """
    labels = [item["label"] for item in batch_items]
    msgs = _build_batch_messages(batch_items)
    raw = await _async_call_api_raw(msgs, semaphore=semaphore)
    if raw is None:
        return {lbl: None for lbl in labels}
    parsed = _parse_batch_response(raw, labels)
    # Sanitize each result
    return {lbl: (_sanitize(text) if text else None) for lbl, text in parsed.items()}


# ---------------------------------------------------------------------------
# TM Index
# ---------------------------------------------------------------------------

class EmbeddingIndex:
    """Semantic TM index using paraphrase-multilingual-MiniLM-L12-v2."""
    MODEL_NAME = "paraphrase-multilingual-MiniLM-L12-v2"

    def __init__(self, entries: list, cache_path: Path = None):
        import numpy as np
        import threading
        # Force offline — model must already be cached locally
        os.environ.setdefault("HF_HUB_OFFLINE", "1")
        os.environ.setdefault("TRANSFORMERS_OFFLINE", "1")
        self.entries = entries
        self._np = np
        self._query_cache: dict = {}  # text -> encoded vector
        self._model_lock = threading.Lock()  # prevent concurrent model init

        # Check whether the embedding cache is valid before loading the model
        cached = None
        if cache_path and cache_path.exists():
            cached = np.load(str(cache_path))   # load once; validate shape
        cache_valid = cached is not None and cached.shape[0] == len(entries)

        if cache_valid:
            self._matrix = cached
            print(f"  Loaded embeddings from cache ({len(entries)} entries, dim={self._matrix.shape[1]})")
            # Load model eagerly even on cache hit — lazy loading causes
            # NotImplementedError under concurrent threads (torch meta tensor bug).
            from sentence_transformers import SentenceTransformer
            print(f"  Loading embedding model ({self.MODEL_NAME})...", end="", flush=True)
            self._model = SentenceTransformer(self.MODEL_NAME)
            print(" done.")
        else:
            from sentence_transformers import SentenceTransformer
            print(f"  Loading embedding model ({self.MODEL_NAME})...", end="", flush=True)
            self._model = SentenceTransformer(self.MODEL_NAME)
            print(" done.")
            if cached is not None:
                print(f"  Cache outdated or invalid — re-encoding.")
            print(f"  Encoding {len(entries)} TM entries...", end="", flush=True)
            self._matrix = self._model.encode(
                [st for st, _ in entries],
                batch_size=256, normalize_embeddings=True,
                show_progress_bar=False, convert_to_numpy=True,
            )
            print(" done.")
            if cache_path:
                np.save(str(cache_path), self._matrix)
                print(f"  Embeddings cached -> {cache_path.name}")

    def _get_model(self):
        """Lazy-load model on first query (cache-hit path skips model load at init)."""
        if self._model is None:
            from sentence_transformers import SentenceTransformer
            print(f"  Loading embedding model ({self.MODEL_NAME})...", end="", flush=True)
            self._model = SentenceTransformer(self.MODEL_NAME)
            print(" done.")
        return self._model

    def query(self, text: str, top_k: int = 3, threshold: float = 0.55) -> list:
        if not text or not text.strip():
            return []
        # Short strings (names, UI labels) need stricter matching to avoid false positives
        effective_threshold = 0.90 if len(text.strip()) < 10 else threshold
        # Cache query vectors to avoid re-encoding the same source text
        vec = self._query_cache.get(text)
        if vec is None:
            vec = self._get_model().encode(
                [text], normalize_embeddings=True, convert_to_numpy=True
            )[0]
            self._query_cache[text] = vec
        sims = self._matrix.dot(vec)
        k = min(top_k, len(sims))
        top_idx = self._np.argpartition(sims, -k)[-k:].tolist()
        results = [(s, self.entries[i][0], self.entries[i][1])
                   for i in top_idx if (s := float(sims[i])) >= effective_threshold]
        results.sort(reverse=True)
        return results


_ONOMATOPOEIA_RE = re.compile(
    r'^[\u54c8\u5475\u5477\u5440\u554a\u554f\u5450\u55ef\u5495\u5561\u5594'
    r'\u5567\u55b5\u5578\u5453\u563f\u5598\u5583\u54e6\u54a7\u54b3\u54c1'
    r'\u54f3\u5514\u5609\u548f\u5555\u5591\u5587\u4e0c\u5501\u55c5\u5600'
    r'\s\?\!\.\u2026]{1,8}$',
    re.UNICODE,
)


def _is_onomatopoeia(st: str) -> bool:
    """Return True if source text is primarily an onomatopoeia or interjection."""
    return bool(_ONOMATOPOEIA_RE.match(st.strip()))


def build_tm_index(entries: list, source_path: str = None) -> tuple:
    """
    Build EmbeddingIndex. Raises if sentence_transformers is unavailable.
    source_path: original TM file path — used to derive .npy cache alongside it.
    Returns (index, description_str).
    """
    cache_path = None
    if source_path:
        p = Path(source_path)
        cache_path = p.parent / (p.stem + ".embeddings.npy")

    idx = EmbeddingIndex(entries, cache_path=cache_path)
    return idx, f"语义向量索引（dim={idx._matrix.shape[1]}）"


# ---------------------------------------------------------------------------
# File / cell utilities
# ---------------------------------------------------------------------------

def get_cell_text(cell) -> str:
    """Return plain-text string from a cell (handles CellRichText)."""
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    try:
        parts = [item.text if hasattr(item, "text") else str(item) for item in val]
        return "".join(parts)
    except TypeError:
        return str(val)


def get_file_headers(filepath: str) -> list:
    """
    Read the first non-empty row of the first sheet.
    Returns list of (1-based col_idx, header_name).
    Supports .xlsx and .csv.
    """
    path = Path(filepath)
    if path.suffix.lower() == ".csv":
        import csv
        with open(filepath, encoding="utf-8-sig", errors="replace") as f:
            row = next(csv.reader(f), [])
        return [(i + 1, v.strip() or f"列{i+1}") for i, v in enumerate(row) if v.strip()]

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(max_row=3, values_only=True):
        non_empty = [(i, v) for i, v in enumerate(row) if v is not None]
        if non_empty:
            last = non_empty[-1][0]
            headers = [(i + 1, str(row[i]) if row[i] is not None else f"列{i+1}")
                       for i in range(last + 1)]
            break
    wb.close()
    return headers


def load_bilingual_file(filepath: str, st_col: int, tt_col: int) -> list:
    """
    Load (st, tt) pairs from xlsx or csv. Used for both glossary and TM.
    Returns list of (st, tt) tuples with empty rows skipped.
    """
    path = Path(filepath)
    rows = []
    if path.suffix.lower() == ".csv":
        import csv
        with open(filepath, encoding="utf-8-sig", errors="replace") as f:
            for row in csv.reader(f):
                if len(row) >= max(st_col, tt_col):
                    st = row[st_col - 1].strip()
                    tt = row[tt_col - 1].strip()
                    if st and tt:
                        rows.append((st, tt))
    else:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        for row in wb.active.iter_rows(values_only=True):
            if row and len(row) >= max(st_col, tt_col):
                st = str(row[st_col - 1] or "").strip()
                tt = str(row[tt_col - 1] or "").strip()
                if st and tt:
                    rows.append((st, tt))
        wb.close()
    return rows


def find_matching_terms(text: str, glossary: list) -> list:
    """Return all (st, tt) pairs whose st appears in text."""
    if not text or not glossary:
        return []
    return [(st, tt) for st, tt in glossary if st in text]


# ---------------------------------------------------------------------------
# Bold markup helpers
# ---------------------------------------------------------------------------

def _diff_bold(original: str, edited: str):
    orig_tokens = re.split(r'(\s+)', original)
    edit_tokens = re.split(r'(\s+)', edited)
    sm = difflib.SequenceMatcher(None, orig_tokens, edit_tokens, autojunk=False)
    blocks, has_change = [], False
    for tag, _, _, j1, j2 in sm.get_opcodes():
        chunk = ''.join(edit_tokens[j1:j2])
        if not chunk:
            continue
        if tag == 'equal':
            blocks.append(chunk)
        else:
            has_change = True
            blocks.append(TextBlock(InlineFont(b=True), chunk))
    return CellRichText(blocks) if has_change else edited


def parse_bold_markup(result: str, original: str = ""):
    """
    Convert [[B]]...[[/B]] markers to CellRichText bold.
    Falls back to word-level diff if no markers but text changed.
    """
    if not result:
        return result
    if '[[B]]' in result:
        if not RICH_TEXT_AVAILABLE:
            return re.sub(r'\[\[/?B\]\]', '', result)
        parts = re.split(r'(\[\[B\]\].*?\[\[/B\]\])', result, flags=re.DOTALL)
        blocks = []
        for part in parts:
            if not part:
                continue
            m = re.match(r'\[\[B\]\](.*?)\[\[/B\]\]', part, re.DOTALL)
            if m:
                blocks.append(TextBlock(InlineFont(b=True), m.group(1)))
            else:
                blocks.append(part)
        return CellRichText(blocks) if blocks else result
    if RICH_TEXT_AVAILABLE and original and original.strip() != result.strip():
        return _diff_bold(original, result)
    return result


def best_tt_variant(tt_variants: list, result: str) -> str:
    """Return the TT variant most similar to result (for bold-diff reference)."""
    if not tt_variants:
        return ""
    if len(tt_variants) == 1:
        return tt_variants[0]
    return max(tt_variants,
               key=lambda v: difflib.SequenceMatcher(None, v, result).ratio())


def _merge(target: dict, source: dict):
    """Merge source counts into target in-place."""
    for k, v in source.items():
        target[k] = target.get(k, 0) + v


def _shift(c, threshold: int, strict: bool = True) -> int:
    """Increment column index c if it is at or above threshold."""
    if not c:
        return c
    return c + 1 if (c > threshold if strict else c >= threshold) else c


# ---------------------------------------------------------------------------
# Sheet processing
# ---------------------------------------------------------------------------

def _cluster_by_st(rows: list) -> list:
    """Group rows with identical ST into clusters (preserving first-appearance order)."""
    st_to_rows = OrderedDict()
    solo = []
    for row in rows:
        st = row[2]
        if not st.strip():
            solo.append([row])
        else:
            st_to_rows.setdefault(st, []).append(row)

    emitted, clusters, solo_idx = set(), [], 0
    for row in rows:
        st = row[2]
        if not st.strip():
            clusters.append(solo[solo_idx])
            solo_idx += 1
        elif st not in emitted:
            clusters.append(st_to_rows[st])
            emitted.add(st)
    return clusters


def process_sheet(sheet, sheet_name: str = "", rows_override: list = None,
                  col_config: dict = None, glossary: list = None,
                  tm_index=None, checkpoint_callback=None) -> tuple:
    """
    Process one sheet: collect rows, cluster by ST, call API concurrently, write results.
    checkpoint_callback: tuple of (_save_workbook, wb, out_path) — saved after every cluster.
    Returns (processed_count, failed_rows, term_counts, row_term_hits).
    """
    ref_cols = (col_config or {}).get("ref_cols", [])
    tm_ref_col = None

    if rows_override is not None:
        rows_to_process = [(r, e, s, k) for r, e, s, k, _ in rows_override]
        post_edit_col = rows_override[0][4]
        if tm_index:
            tm_ref_col = post_edit_col + 1
        total = len(rows_to_process)
    else:
        tt_col  = col_config["tt"]
        st_col  = col_config.get("st")
        key_col = col_config.get("key")

        # Find header row
        header_row = next(
            (r for r in range(1, 4) if sheet.cell(row=r, column=tt_col).value is not None),
            1
        )

        post_edit_col = tt_col + 1
        resuming = sheet.cell(row=header_row, column=post_edit_col).value == config.OUTPUT_COLUMN_HEADER

        if resuming:
            print(f"    [RESUME] '{config.OUTPUT_COLUMN_HEADER}' column found — skipping filled rows.")
            if tm_index:
                if sheet.cell(row=header_row, column=post_edit_col + 1).value == "TM Reference":
                    tm_ref_col = post_edit_col + 1
        else:
            sheet.insert_cols(post_edit_col)
            # Adjust column indices after insert
            st_col  = _shift(st_col,  tt_col, strict=True)
            key_col = _shift(key_col, tt_col, strict=True)
            for rc in ref_cols:
                if rc["col"] > tt_col:
                    rc["col"] += 1
            sheet.cell(row=header_row, column=post_edit_col).value = config.OUTPUT_COLUMN_HEADER

            if tm_index:
                tm_ref_col = post_edit_col + 1
                sheet.insert_cols(tm_ref_col)
                st_col  = _shift(st_col,  tm_ref_col, strict=False)
                key_col = _shift(key_col, tm_ref_col, strict=False)
                for rc in ref_cols:
                    if rc["col"] >= tm_ref_col:
                        rc["col"] += 1
                sheet.cell(row=header_row, column=tm_ref_col).value = "TM Reference"

        rows_to_process = []
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            tt = get_cell_text(sheet.cell(row=row_idx, column=tt_col))
            if not tt.strip():
                continue
            if resuming and sheet.cell(row=row_idx, column=post_edit_col).value:
                continue
            st  = get_cell_text(sheet.cell(row=row_idx, column=st_col))  if st_col  else ""
            key = get_cell_text(sheet.cell(row=row_idx, column=key_col)) if key_col else ""
            rows_to_process.append((row_idx, tt, st, key))

        total = len(rows_to_process)

    if total == 0:
        print("    [OK] Nothing to process.")
        return 0, [], {}, {}

    clusters = _cluster_by_st(rows_to_process)
    multi = sum(1 for c in clusters if len(c) > 1)
    if multi:
        print(f"    [Cluster] {total} rows → {len(clusters)} clusters ({multi} multi-row)")

    # Unpack checkpoint params (used by the async writer)
    if checkpoint_callback:
        _save_fn, _ckpt_wb, _ckpt_path = checkpoint_callback
    else:
        _save_fn = _ckpt_wb = _ckpt_path = None

    # ------------------------------------------------------------------
    # Async inner function: packs clusters into batches, one API call
    # per batch. Rows within a batch share context naturally.
    # ------------------------------------------------------------------
    async def _run_async():
        semaphore   = asyncio.Semaphore(config.MAX_WORKERS)
        completed   = [0]
        term_counts : dict = {}
        row_term_hits: dict = {}
        failed_rows : list = []
        processed   = [0]
        _last_save  = [0.0]
        _SAVE_INTERVAL = 5.0
        loop = asyncio.get_running_loop()

        # Step 1: pre-compute per-cluster metadata
        batch_size = max(1, getattr(config, "MTPE_BATCH_SIZE", 15))

        cluster_meta = []
        for cluster in clusters:
            st_text = cluster[0][2]

            seen, tt_variants, key_texts = set(), [], []
            for _, tt, _, key in cluster:
                if tt not in seen:
                    seen.add(tt)
                    tt_variants.append(tt)
                    key_texts.append(key)
            tt_best  = tt_variants[0]
            key_best = key_texts[0] if key_texts else ""

            hints = find_matching_terms(st_text, glossary) if glossary else []
            if hints:
                for pair in hints:
                    term_counts[pair] = term_counts.get(pair, 0) + 1

            ref_data = [
                {"note": rc["note"], "value": v}
                for rc in ref_cols
                if (v := get_cell_text(sheet.cell(row=cluster[0][0], column=rc["col"]))).strip()
            ]

            cluster_meta.append((cluster, st_text, tt_best, key_best, hints, ref_data))

        # Step 2: TM queries concurrently
        async def _tm_query(st_text):
            if tm_index and st_text.strip() and not _is_onomatopoeia(st_text):
                return await loop.run_in_executor(None, tm_index.query, st_text)
            return []

        tm_hits_list = await asyncio.gather(*[_tm_query(m[1]) for m in cluster_meta])

        # Step 3: pack clusters into batches
        batches = []
        current_batch = []
        label_counter = 1

        for meta, tm_hits in zip(cluster_meta, tm_hits_list):
            cluster, st_text, tt_best, key_best, hints, ref_data = meta
            item = {
                "label":    label_counter,
                "st":       st_text,
                "tt":       tt_best,
                "key":      key_best,
                "hints":    hints,
                "tm_hits":  tm_hits,
                "ref_data": ref_data,
                "cluster":  cluster,
            }
            current_batch.append(item)
            label_counter += 1
            if len(current_batch) >= batch_size:
                batches.append(current_batch)
                current_batch = []
                label_counter = 1

        if current_batch:
            batches.append(current_batch)

        # Step 4: dispatch batches concurrently
        async def process_batch_async(batch):
            results_dict = await _async_call_api_batch(batch, semaphore=semaphore)
            completed[0] += sum(len(item["cluster"]) for item in batch)
            print("[%d/%d] batch of %d rows done" % (completed[0], total, len(batch)), end=""  , flush=True)
            return batch, results_dict

        tasks = [asyncio.create_task(process_batch_async(b)) for b in batches]

        for coro in asyncio.as_completed(tasks):
            batch, results_dict = await coro

            for item in batch:
                lbl      = item["label"]
                cluster  = item["cluster"]
                tt_best  = item["tt"]
                tm_hits  = item["tm_hits"]
                hints    = item["hints"]
                result   = results_dict.get(lbl)

                tm_str = "\n".join(
                    "[%.0f%%] %s -> %s" % (s*100, st_tm, tt_tm)
                    for s, st_tm, tt_tm in tm_hits
                ) if tm_ref_col and tm_hits else None
                for row in cluster:
                    row_idx, tt_text = row[0], row[1]
                    if result is None:
                        failed_rows.append((row_idx, tt_text, row[2], row[3], post_edit_col))
                    else:
                        cell_val = parse_bold_markup(result, original=tt_best)
                        sheet.cell(row=row_idx, column=post_edit_col).value = cell_val
                        if tm_str is not None:
                            sheet.cell(row=row_idx, column=tm_ref_col).value = tm_str
                        if hints:
                            row_term_hits[row_idx] = hints
                        processed[0] += 1

            if _save_fn:
                now = loop.time()
                if now - _last_save[0] >= _SAVE_INTERVAL:
                    _last_save[0] = now
                    try:
                        await loop.run_in_executor(None, _save_fn, _ckpt_wb, _ckpt_path)
                    except Exception as e:
                        print("\r    [Checkpoint] Save failed:", e, flush=True)

        if _save_fn:
            try:
                await loop.run_in_executor(None, _save_fn, _ckpt_wb, _ckpt_path)
            except Exception as e:
                print("\r    [Checkpoint] Final save failed:", e, flush=True)

        return processed[0], failed_rows, term_counts, row_term_hits
    # Run the async event loop (blocks the calling thread until done)
    processed, failed_rows, term_counts, row_term_hits = asyncio.run(_run_async())

    status = f"{processed}/{total} row(s) processed"
    if failed_rows:
        status += f", {len(failed_rows)} failed"
    print(f"\r    Done — {status}.{' ' * 50}")
    return processed, failed_rows, term_counts, row_term_hits


# ---------------------------------------------------------------------------
# File processing
# ---------------------------------------------------------------------------

def _save_workbook(wb, path: Path) -> Path:
    """Save workbook; on PermissionError append timestamp to filename."""
    try:
        wb.save(str(path))
        return path
    except PermissionError:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = path.parent / (path.stem + f"_{ts}" + path.suffix)
        wb.save(str(fallback))
        print(f"\n  [WARNING] File locked — saved to {fallback.name}")
        return fallback


def process_file(filepath: str, col_config: dict = None,
                 glossary: list = None, tm_index=None,
                 retry_prompt=None) -> tuple:
    """
    Load workbook, process every sheet, save as *_postedit.xlsx.
    Every completed cluster is immediately saved to disk (per-row checkpoint).
    retry_prompt: callable(msg) -> str for interactive retry; if None, auto-skip failed rows.
    Returns (success: bool, term_counts: dict, row_term_hits: dict).
    row_term_hits maps row_idx -> [(st, tt)] for terminology hits recorded
    during MTPE, to be passed to qa_and_repair for precise QA context.
    """
    path = Path(filepath)
    print(f"\n{'='*60}\n  File: {path.name}\n{'='*60}")

    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.OUTPUT_DIR / (path.stem + config.OUTPUT_SUFFIX + path.suffix)

    # Load from output file if it exists (resume), else load from source
    load_path = out_path if out_path.exists() else path
    try:
        wb = load_workbook(str(load_path), rich_text=True)
    except Exception as exc:
        print(f"  [ERROR] Could not open file: {exc}")
        return False, {}, {}

    if load_path == out_path:
        print(f"  [RESUME] Loading from existing output file.")

    all_failed, file_term_counts, file_row_term_hits = {}, {}, {}

    for sheet_name in wb.sheetnames:
        print(f"\n  Sheet: {sheet_name}")
        processed, failed, tc, rth = process_sheet(
            wb[sheet_name], sheet_name,
            col_config=col_config, glossary=glossary, tm_index=tm_index,
            checkpoint_callback=(_save_workbook, wb, out_path),
        )
        _merge(file_term_counts, tc)
        file_row_term_hits.update(rth)
        if failed:
            all_failed[sheet_name] = failed

    # Auto-retry failed rows (no interaction needed) — covers 429 transient failures
    if all_failed:
        total_failed = sum(len(v) for v in all_failed.values())
        print(f"\n  Auto-retrying {total_failed} failed rows...")
        new_failed = {}
        for sheet_name, rows in all_failed.items():
            print(f"\n  Sheet: {sheet_name} (retrying {len(rows)} rows)")
            _, failed, tc, rth = process_sheet(
                wb[sheet_name], sheet_name, rows_override=rows,
                col_config=col_config, glossary=glossary, tm_index=tm_index,
                checkpoint_callback=(_save_workbook, wb, out_path),
            )
            _merge(file_term_counts, tc)
            file_row_term_hits.update(rth)
            if failed:
                new_failed[sheet_name] = failed
        all_failed = new_failed

    # Interactive retry loop (used by main.py CLI)
    while all_failed and retry_prompt:
        total_failed = sum(len(v) for v in all_failed.values())
        print(f"\n  {total_failed} rows failed:")
        for sheet_name, rows in all_failed.items():
            for row_idx, en_text, _, key_text, _ in rows:
                key_hint = f"[{key_text[:30]}] " if key_text else ""
                print(f"    Sheet '{sheet_name}'  Row {row_idx:>4}  {key_hint}{en_text[:45]}")
        print("\n  Retry?\n    1. Yes\n    2. No, skip and save")
        if retry_prompt("Please select") != "1":
            break
        new_failed = {}
        for sheet_name, rows in all_failed.items():
            print(f"\n  Sheet: {sheet_name} (retrying {len(rows)} rows)")
            _, failed, tc, rth = process_sheet(
                wb[sheet_name], sheet_name, rows_override=rows,
                col_config=col_config, glossary=glossary, tm_index=tm_index
            )
            _merge(file_term_counts, tc)
            file_row_term_hits.update(rth)
            if failed:
                new_failed[sheet_name] = failed
        all_failed = new_failed

    saved = _save_workbook(wb, out_path)
    print(f"\n  Saved -> {saved.name}")
    return True, file_term_counts, file_row_term_hits
